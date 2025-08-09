#!/usr/bin/env python3
"""
Simple Schedule Generator
Creates a basic working schedule using minimal constraints
"""

import json
import os
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

def load_data():
    """Load event parameters and constraints"""
    try:
        with open('schedule/event_par.txt', 'r', encoding='utf-8') as f:
            event_params = json.load(f)
        
        with open('schedule/constraints.txt', 'r', encoding='utf-8') as f:
            constraints_data = json.load(f)
        
        return event_params, constraints_data
    except Exception as e:
        print(f"❌ Error loading data: {e}")
        return None, None

def generate_simple_schedule():
    """Generate a simple working schedule"""
    print("🚀 Simple Schedule Generator")
    print("="*50)
    
    event_params, constraints_data = load_data()
    if not event_params or not constraints_data:
        return False
    
    start_date = datetime.strptime(event_params['start_date'], '%Y-%m-%d').date()
    end_date = datetime.strptime(event_params['end_date'], '%Y-%m-%d').date()
    employees = constraints_data.get('employees', [])
    
    print(f"📅 Period: {start_date} to {end_date}")
    print(f"👥 Soldiers: {len(employees)}")
    
    # Convert constraints to date objects
    for emp in employees:
        emp['unavailable_dates'] = []
        for date_str in emp.get('unavailableDays', []):
            try:
                constraint_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                if start_date <= constraint_date <= end_date:
                    emp['unavailable_dates'].append(constraint_date)
            except ValueError:
                continue
    
    # Generate calendar
    total_days = (end_date - start_date).days + 1
    dates = [start_date + timedelta(days=i) for i in range(total_days)]
    
    # Simple rotation algorithm
    schedule = {}
    soldiers_per_day = 4  # Reduced requirement
    
    for emp in employees:
        schedule[emp['name']] = {'schedule': [], 'base_days': 0, 'home_days': 0}
    
    # Create rotation schedule
    available_soldiers = [emp['name'] for emp in employees]
    current_rotation = 0
    
    for current_date in dates:
        is_weekend = current_date.weekday() == 5  # Saturday
        daily_assignments = []
        
        # Get available soldiers for this date
        available_today = []
        for emp in employees:
            if current_date not in emp['unavailable_dates']:
                # Special handling for weekend soldier ינון
                if emp['name'] == 'ינון':
                    if is_weekend:
                        available_today.append(emp['name'])
                else:
                    available_today.append(emp['name'])
        
        # Assign soldiers using simple rotation
        assigned_count = 0
        rotation_start = current_rotation
        
        while assigned_count < soldiers_per_day and len(available_today) > 0:
            if rotation_start >= len(available_today):
                rotation_start = 0
            
            if rotation_start < len(available_today):
                soldier = available_today[rotation_start]
                if soldier not in daily_assignments:
                    daily_assignments.append(soldier)
                    assigned_count += 1
            
            rotation_start += 1
            if rotation_start >= len(available_today) and assigned_count < soldiers_per_day:
                break
        
        # Update rotation counter
        current_rotation = (current_rotation + soldiers_per_day) % len(employees) if employees else 0
        
        # Record assignments
        for emp in employees:
            is_on_base = emp['name'] in daily_assignments
            status = 'Base' if is_on_base else 'Home'
            
            schedule[emp['name']]['schedule'].append({
                'date': current_date.isoformat(),
                'status': status,
                'is_weekend': is_weekend
            })
            
            if is_on_base:
                schedule[emp['name']]['base_days'] += 1
            else:
                schedule[emp['name']]['home_days'] += 1
    
    return schedule, employees, dates

def export_to_json(schedule, filename="simple_schedule_results.json"):
    """Export schedule to JSON"""
    print(f"📄 Exporting to {filename}...")
    
    export_data = {
        'metadata': {
            'generated_at': datetime.now().isoformat(),
            'generator': 'simple_schedule_generator',
            'total_soldiers': len(schedule)
        },
        'soldiers': {}
    }
    
    for soldier_name, soldier_data in schedule.items():
        export_data['soldiers'][soldier_name] = {
            'soldier_info': {
                'name': soldier_name,
                'is_weekend_only': soldier_name == 'ינון'
            },
            'schedule': soldier_data['schedule'],
            'summary': {
                'total_base_days': soldier_data['base_days'],
                'total_home_days': soldier_data['home_days']
            }
        }
    
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, indent=2, ensure_ascii=False)
    
    print(f"✅ JSON export completed: {filename}")

def export_to_excel(schedule, employees, dates, filename="simple_schedule_results.xlsx"):
    """Export schedule to Excel"""
    print(f"📊 Exporting to {filename}...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    
    # Colors
    base_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")  # Green
    home_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")  # Amber
    weekend_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")  # Blue
    
    # Headers
    ws['A1'] = 'Soldier'
    ws['A1'].font = Font(bold=True)
    
    # Date headers
    for col, date_obj in enumerate(dates, start=2):
        cell = ws.cell(row=1, column=col)
        cell.value = date_obj.strftime('%m/%d')
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
        if date_obj.weekday() == 5:  # Saturday
            cell.fill = weekend_fill
    
    # Soldier data
    for row, emp in enumerate(employees, start=2):
        ws.cell(row=row, column=1, value=emp['name'])
        soldier_schedule = schedule[emp['name']]['schedule']
        
        for col, day_data in enumerate(soldier_schedule, start=2):
            cell = ws.cell(row=row, column=col)
            cell.value = day_data['status']
            cell.alignment = Alignment(horizontal='center')
            
            if day_data['status'] == 'Base':
                if day_data['is_weekend']:
                    cell.fill = weekend_fill
                else:
                    cell.fill = base_fill
            else:
                cell.fill = home_fill
    
    # Summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws['A1'] = 'Soldier'
    summary_ws['B1'] = 'Base Days'
    summary_ws['C1'] = 'Home Days'
    summary_ws['D1'] = 'Total Days'
    
    for col in range(1, 5):
        summary_ws.cell(row=1, column=col).font = Font(bold=True)
    
    for row, emp in enumerate(employees, start=2):
        soldier_data = schedule[emp['name']]
        summary_ws.cell(row=row, column=1, value=emp['name'])
        summary_ws.cell(row=row, column=2, value=soldier_data['base_days'])
        summary_ws.cell(row=row, column=3, value=soldier_data['home_days'])
        summary_ws.cell(row=row, column=4, value=soldier_data['base_days'] + soldier_data['home_days'])
    
    wb.save(filename)
    print(f"✅ Excel export completed: {filename}")

def main():
    """Main function"""
    try:
        schedule, employees, dates = generate_simple_schedule()
        if not schedule:
            return False
        
        # Export results
        export_to_json(schedule)
        export_to_excel(schedule, employees, dates)
        
        print("\n" + "="*50)
        print("✅ SIMPLE SCHEDULE GENERATION COMPLETED")
        print("📁 Generated Files:")
        print("   • simple_schedule_results.json")
        print("   • simple_schedule_results.xlsx")
        print("="*50)
        
        return True
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = main()
    exit(0 if success else 1)