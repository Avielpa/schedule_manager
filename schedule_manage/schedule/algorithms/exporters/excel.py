"""
📊 Excel Exporter - V9.0
Exports scheduling results to Excel with advanced formatting and summaries
"""

import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import date
from typing import List, Dict, Any

from ..config import (
    COLOR_BASE, COLOR_HOME, COLOR_WEEKEND, COLOR_HEADER, COLOR_SUMMARY,
    EXCEL_COLUMN_WIDTH_NAME, EXCEL_COLUMN_WIDTH_DATE, EXCEL_COLUMN_WIDTH_SUMMARY
)
from ..utils import format_date_for_excel, is_saturday


class ExcelExporter:
    """Class for exporting results to Excel"""
    
    def __init__(self, calendar: List[date], soldiers: List):
        self.calendar = calendar
        self.soldiers = soldiers
        self._setup_styles()
    
    def _setup_styles(self) -> None:
        """Sets up the styles for Excel"""
        # Borders
        self.thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Text alignment
        self.center_aligned_text = Alignment(horizontal="center", vertical="center")
        
        # Fonts
        self.bold_font = Font(bold=True)
        self.title_font = Font(bold=True, size=14)
        self.exceptional_font = Font(bold=True, color="FF6600")  # Orange for exceptional soldiers
        self.weekend_soldier_font = Font(bold=True, color="0066FF")  # Blue for weekend soldiers
        
        # Background colors
        self.header_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        self.base_fill = PatternFill(start_color=COLOR_BASE, end_color=COLOR_BASE, fill_type="solid")
        self.home_fill = PatternFill(start_color=COLOR_HOME, end_color=COLOR_HOME, fill_type="solid")
        self.weekend_base_fill = PatternFill(start_color=COLOR_WEEKEND, end_color=COLOR_WEEKEND, fill_type="solid")
        self.summary_fill = PatternFill(start_color=COLOR_SUMMARY, end_color=COLOR_SUMMARY, fill_type="solid")
    
    def export(self, solution_data: Dict[str, Any], output_filepath: str) -> None:
        """
        Exports the solution to Excel
        
        Args:
            solution_data: The solution data
            output_filepath: The output file path
        """
        if not solution_data:
            print("❌ No solution data to export to Excel")
            return
        
        print(f"📝 Exporting schedule to Excel: {output_filepath}")
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Schedule V9.0"
            
            self._create_header(ws)
            self._create_schedule_table(ws, solution_data)
            self._create_daily_summary(ws, solution_data)
            self._adjust_column_widths(ws)
            
            wb.save(output_filepath)
            print("✅ Excel export completed successfully")
            
        except Exception as e:
            print(f"❌ Error exporting to Excel: {e}")
    
    def _create_header(self, ws) -> None:
        """Creates the main header"""
        # Main title
        ws['A1'] = "💎 DynamicBlockCrusher V9.0 - Advanced Schedule"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.center_aligned_text
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self.calendar) + 4)
        
        # Capabilities description
        ws['A2'] = "🎯 Personal Targets | 🏖️ Adapted Weekends | 🔧 Intelligent Softening | 🚨 One-Day Block Prevention | ⚖️ Balance and Fairness"
        ws['A2'].font = Font(size=10, italic=True)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(self.calendar) + 4)
    
    def _create_schedule_table(self, ws, solution_data: Dict) -> None:
        """Creates the main schedule table"""
        start_row = 4
        
        # Column headers
        self._create_column_headers(ws, start_row)
        
        # Fill soldier data
        current_row = start_row + 1
        sorted_soldiers = sorted(self.soldiers, key=lambda s: s.name)
        
        for soldier in sorted_soldiers:
            self._create_soldier_row(ws, current_row, soldier, solution_data)
            current_row += 1
    
    def _create_column_headers(self, ws, start_row: int) -> None:
        """Creates the column headers"""
        current_col = 1
        
        # Soldier names column
        cell = ws.cell(row=start_row, column=current_col, value="Soldier Name")
        self._apply_header_style(cell)
        current_col += 1
        
        # Date columns
        for day in self.calendar:
            day_str = format_date_for_excel(day)
            cell = ws.cell(row=start_row, column=current_col, value=day_str)
            self._apply_header_style(cell)
            current_col += 1
        
        # Summary columns
        summary_headers = ["Base Days", "Home Days", "Weekends"]
        for header in summary_headers:
            cell = ws.cell(row=start_row, column=current_col, value=header)
            self._apply_header_style(cell)
            current_col += 1
    
    def _create_soldier_row(self, ws, row: int, soldier, solution_data: Dict) -> None:
        """Creates a row for a specific soldier"""
        # Soldier name with emphasis for different types
        self._create_soldier_name_cell(ws, row, soldier)
        
        # Schedule days
        self._create_soldier_schedule_cells(ws, row, soldier, solution_data)
        
        # Summaries
        self._create_soldier_summary_cells(ws, row, soldier, solution_data)
    
    def _create_soldier_name_cell(self, ws, row: int, soldier) -> None:
        """Creates a cell with the soldier's name"""
        cell = ws.cell(row=row, column=1)
        
        constraints_ratio = len(soldier.raw_constraints) / len(self.calendar)
        
        if soldier.is_weekend_only_soldier_flag:
            cell.value = f"🏖️ {soldier.name}"
            cell.font = self.weekend_soldier_font
        elif soldier.is_exceptional_output or constraints_ratio > 0.4:
            cell.value = f"⚠️ {soldier.name}"
            cell.font = self.exceptional_font
        else:
            cell.value = soldier.name
            cell.font = self.bold_font
        
        cell.border = self.thin_border
        cell.alignment = self.center_aligned_text
    
    def _create_soldier_schedule_cells(self, ws, row: int, soldier, solution_data: Dict) -> None:
        """Creates daily schedule cells for the soldier"""
        if soldier.name not in solution_data:
            return
        
        schedule = solution_data[soldier.name]['schedule']
        constraints_ratio = len(soldier.raw_constraints) / len(self.calendar)
        
        for col_idx, day_data in enumerate(schedule):
            col = col_idx + 2  # Start from column 2
            cell = ws.cell(row=row, column=col)
            cell.alignment = self.center_aligned_text
            cell.border = self.thin_border
            
            current_date = date.fromisoformat(day_data['date'])
            status = day_data['status']
            
            if status == 'Base':
                # Different symbols for soldier types
                if soldier.is_weekend_only_soldier_flag:
                    cell.value = "🏖️"
                elif constraints_ratio > 0.4:
                    cell.value = "⚠️"
                else:
                    cell.value = "🪖"
                
                # Background color
                if is_saturday(current_date):
                    cell.fill = self.weekend_base_fill
                else:
                    cell.fill = self.base_fill
            else:
                cell.value = "🏠"
                cell.fill = self.home_fill
    
    def _create_soldier_summary_cells(self, ws, row: int, soldier, solution_data: Dict) -> None:
        """Creates summary cells for the soldier"""
        if soldier.name not in solution_data:
            return
        
        soldier_data = solution_data[soldier.name]
        summary_col_start = len(self.calendar) + 2
        
        summary_values = [
            soldier_data.get('total_base_days', 0),
            soldier_data.get('total_home_days', 0),
            soldier_data.get('total_weekend_base_days', 0)
        ]
        
        for i, value in enumerate(summary_values):
            cell = ws.cell(row=row, column=summary_col_start + i, value=value)
            cell.alignment = self.center_aligned_text
            cell.border = self.thin_border
            cell.fill = self.summary_fill
    
    def _create_daily_summary(self, ws, solution_data: Dict) -> None:
        """Creates daily summary row"""
        soldiers_count = len(self.soldiers)
        summary_row = 4 + soldiers_count + 2  # Below the main table
        
        # Summary row title
        cell = ws.cell(row=summary_row, column=1, value="Total Soldiers on Base")
        cell.font = self.bold_font
        cell.alignment = self.center_aligned_text
        cell.fill = self.header_fill
        cell.border = self.thin_border
        
        # Fill daily count
        if 'daily_soldiers_count' in solution_data:
            for col_idx, day in enumerate(self.calendar):
                col = col_idx + 2
                count = solution_data['daily_soldiers_count'].get(day.isoformat(), 0)
                
                cell = ws.cell(row=summary_row, column=col, value=count)
                cell.alignment = self.center_aligned_text
                cell.border = self.thin_border
                cell.fill = self.header_fill
                cell.font = self.bold_font
    
    def _apply_header_style(self, cell) -> None:
        """Applies header style to a cell"""
        cell.font = self.bold_font
        cell.alignment = self.center_aligned_text
        cell.fill = self.header_fill
        cell.border = self.thin_border
    
    def _adjust_column_widths(self, ws) -> None:
        """Adjusts column widths"""
        # Name column
        ws.column_dimensions['A'].width = EXCEL_COLUMN_WIDTH_NAME
        
        # Date columns
        for col_idx in range(2, len(self.calendar) + 2):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = EXCEL_COLUMN_WIDTH_DATE
        
        # Summary columns
        summary_start_col = len(self.calendar) + 2
        for col_idx in range(summary_start_col, summary_start_col + 3):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = EXCEL_COLUMN_WIDTH_SUMMARY
    
    def create_advanced_excel_report(self, solution_data: Dict, output_filepath: str, 
                                   include_statistics: bool = True) -> None:
        """
        Creates an advanced Excel report with additional statistics
        
        Args:
            solution_data: The solution data
            output_filepath: The output file path
            include_statistics: Whether to include advanced statistics
        """
        if not solution_data:
            print("❌ No solution data for advanced report")
            return
        
        print(f"📊 Creating advanced Excel report: {output_filepath}")
        
        try:
            wb = openpyxl.Workbook()
            
            # Main sheet - Schedule
            ws_main = wb.active
            ws_main.title = "Schedule"
            self._create_main_schedule_sheet(ws_main, solution_data)
            
            if include_statistics:
                # Statistics sheet
                ws_stats = wb.create_sheet("Statistics")
                self._create_statistics_sheet(ws_stats, solution_data)
                
                # Soldier analysis sheet
                ws_analysis = wb.create_sheet("Soldier Analysis")
                self._create_soldiers_analysis_sheet(ws_analysis, solution_data)
            
            wb.save(output_filepath)
            print("✅ Advanced Excel report created successfully")
            
        except Exception as e:
            print(f"❌ Error creating advanced report: {e}")
    
    def _create_main_schedule_sheet(self, ws, solution_data: Dict) -> None:
        """Creates the main sheet"""
        self._create_header(ws)
        self._create_schedule_table(ws, solution_data)
        self._create_daily_summary(ws, solution_data)
        self._adjust_column_widths(ws)
    
    def _create_statistics_sheet(self, ws, solution_data: Dict) -> None:
        """Creates a statistics sheet"""
        ws['A1'] = "📊 General Statistics"
        ws['A1'].font = self.title_font
        
        row = 3
        
        # Basic statistics
        stats = self._calculate_basic_statistics(solution_data)
        for key, value in stats.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            row += 1
    
    def _create_soldiers_analysis_sheet(self, ws, solution_data: Dict) -> None:
        """Creates a soldier analysis sheet"""
        ws['A1'] = "👥 Soldier Analysis"
        ws['A1'].font = self.title_font
        
        # Headers
        headers = ["Soldier Name", "Base Days", "Home Days", "Weekends", "Availability %", "Type"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._apply_header_style(cell)
        
        # Soldier data
        row = 4
        for soldier in self.soldiers:
            if soldier.name in solution_data:
                self._add_soldier_analysis_row(ws, row, soldier, solution_data)
                row += 1
    
    def _calculate_basic_statistics(self, solution_data: Dict) -> Dict[str, Any]:
        """Calculates basic statistics"""
        total_soldiers = len(self.soldiers)
        total_days = len(self.calendar)
        
        total_base_days = sum(
            data.get('total_base_days', 0) 
            for name, data in solution_data.items() 
            if name != 'daily_soldiers_count'
        )
        
        return {
            'Total Soldiers': total_soldiers,
            'Total Days': total_days,
            'Total Base Days': total_base_days,
            'Average Base Days per Soldier': round(total_base_days / total_soldiers, 2),
            'Weekends in Calendar': len([d for d in self.calendar if is_saturday(d)])
        }
    
    def _add_soldier_analysis_row(self, ws, row: int, soldier, solution_data: Dict) -> None:
        """
        Adds a soldier analysis row
        """
        soldier_data = solution_data[soldier.name]
        constraints_ratio = len(soldier.raw_constraints) / len(self.calendar)
        
        data = [
            soldier.name,
            soldier_data.get('total_base_days', 0),
            soldier_data.get('total_home_days', 0),
            soldier_data.get('total_weekend_base_days', 0),
            f"{(1 - constraints_ratio) * 100:.1f}%",
            "Weekend" if soldier.is_weekend_only_soldier_flag else "Regular"
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value).border = self.thin_border